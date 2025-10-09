#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GPI Survey QC Tool — GUI (Excel or HTML report)

Includes:
- Gear (⚙) Settings button: choose Excel (.xlsx) or HTML (.html) output
- Clear help text explaining elevation parameters + recommended ranges
- QC logic per your approved version (unchanged)

Dependencies: pandas, numpy, openpyxl, xlsxwriter
"""

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from pathlib import Path
from datetime import datetime
import threading
import queue
import base64
import pandas as pd
import numpy as np
import re
import sys
import traceback
import uuid

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
except Exception:  # pragma: no cover - optional dependency
    DND_FILES = None
    TkinterDnD = None

# ============ Theme ============
GPI_GREEN = "#0F3320"
GPI_GREY = "#A2AAAD"
GPI_MED = "#427829"
GPI_HL = "#84BD00"
BG = "#F4F6F5"

# ============ Optional styling (safe if not installed) ============
try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font
except Exception:
    load_workbook = None
    PatternFill = None
    Font = None

try:
    import xlsxwriter  # noqa: F401
    XLSX_ENGINE = "xlsxwriter"
except Exception:
    XLSX_ENGINE = "openpyxl"

try:
    from PIL import Image, ImageTk
except Exception:  # pragma: no cover - optional dependency
    Image = None
    ImageTk = None

# ============ Regex / parsing ============
CONTROL_START = {"ST", "START", "BEG", "BEGIN"}
CONTROL_END = {"END", "STOP"}
CONTROL_CLOSE = {"CL", "CLS", "CLOSE", "CLOS"}
CONTROL_PC = {"PC"}
CONTROL_PT = {"PT"}
CTRL_RE = re.compile(
    r"^(ST|START|BEG|BEGIN|END|STOP|CL|CLS|CLOSE|CLOS|PC|PT)[\.,;:]*$",
    re.IGNORECASE,
)
BASE_RE = re.compile(r"^([A-Za-z]+(?:-[A-Za-z]+)*)(\d*)$")

def strip_num_suffix(s: str) -> str:
    if not isinstance(s, str):
        s = "" if pd.isna(s) else str(s)
    token = s.strip().split()[0] if s else ""
    m = BASE_RE.match(token.upper())
    return m.group(1) if m else token.upper()


def _generate_error_code() -> str:
    """Return a short random error code for crash reporting."""
    return f"E{uuid.uuid4().hex[:8].upper()}"


def _write_crash_log(error_code: str, exc: Exception, *, context=None, preferred_dir=None):
    """Persist diagnostic information about a crash.

    Returns the path of the written log file, or ``None`` if logging failed.
    """
    context = context or {}
    timestamp = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    tb = "".join(traceback.format_exception(type(exc), exc, exc.__traceback__))
    lines = [
        "GPI Survey QC Tool crash report",
        f"Timestamp: {timestamp}",
        f"Error code: {error_code}",
        f"Python: {sys.version}",
    ]
    for key, value in context.items():
        lines.append(f"{key}: {value}")
    lines.append("\nTraceback:\n" + tb)

    log_content = "\n".join(lines)
    log_name = f"gpi_qc_tool_{error_code}.log"

    candidate_dirs = []
    if preferred_dir:
        try:
            candidate_dirs.append(Path(preferred_dir))
        except Exception:
            pass
    candidate_dirs.extend([
        Path.home() / "GPI_QC_Tool_Logs",
        Path(__file__).resolve().parent / "logs",
    ])

    for directory in candidate_dirs:
        try:
            directory.mkdir(parents=True, exist_ok=True)
            log_path = directory / log_name
            log_path.write_text(log_content, encoding="utf-8")
            return log_path
        except Exception:
            continue
    return None

def extract_line_events_field(text: str):
    """Stream tokens left->right, pairing each control with the most recent base."""
    if not isinstance(text, str) or not text.strip():
        return []
    parts = re.split(r"[|\s]+", text.strip())
    events = []
    current_base, current_lineid = None, ""
    for p in parts:
        pu = p.strip()
        if not pu:
            continue
        mctrl = CTRL_RE.match(pu.upper())
        if mctrl:
            ctrl = mctrl.group(1).upper()
            events.append((current_base, current_lineid, ctrl))
            continue
        mbase = BASE_RE.match(pu.upper())
        if mbase:
            current_base, current_lineid = mbase.group(1), mbase.group(2)
            continue
    return events

# ============ IO helpers ============
def read_csv_any(path: Path) -> pd.DataFrame:
    for enc in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            return pd.read_csv(path, encoding=enc)
        except Exception:
            continue
    return pd.read_csv(path)  # last resort

def normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip().replace("\ufeff", "") for c in df.columns]
    return df

def best_col(df: pd.DataFrame, canonical: str) -> str:
    cand = canonical.lower()
    aliases = {
        "point id": ["point id", "point", "pointid", "point no", "point number", "pt id", "ptid"],
        "northing": ["northing", "north", "y"],
        "easting": ["easting", "east", "x"],
        "elevation": ["elevation", "elev", "z"],
        "feature code": ["feature code", "feature", "code", "feature_code"],
        "attribute": ["attribute", "attributes", "attr", "remarks", "desc", "description"],
        "alpha code": ["alpha code", "alphacode", "alpha", "code"],
        "attribute type": ["attribute type", "attr type", "attrib type", "type"],
    }
    # Perfect match?
    for c in df.columns:
        if c.lower().strip() == cand:
            return c
    # Alias
    for c in df.columns:
        if c.lower().strip() in aliases.get(cand, []):
            return c
    raise ValueError(f"Missing required column: {canonical}")

# ============ QC core ============
def local_plane_outliers(df: pd.DataFrame,
                         radius=100.0, min_k=8, k_fallback=15, max_k=30,
                         iqr_mult=3.0, mad_mult=6.0) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame({"Info": ["No terrain-eligible points to evaluate"]})
    coords = df[["Easting_num", "Northing_num"]].to_numpy()
    z = df["Elevation_num"].to_numpy()
    valid = ~np.isnan(coords).any(axis=1) & ~np.isnan(z)
    idxs = np.where(valid)[0]
    if len(idxs) < (min_k + 1):
        return pd.DataFrame({"Info": ["Too few valid points for local analysis"]})

    pts = coords[idxs]
    dmat = np.sqrt(((pts[:, None, :] - pts[None, :, :])**2).sum(axis=2))

    def fit_plane(X, y):
        A = np.c_[X[:, 0], X[:, 1], np.ones(len(X))]
        sol, *_ = np.linalg.lstsq(A, y, rcond=None)
        return sol

    out = []
    for ii, gi in enumerate(idxs):
        within_r = np.where((dmat[ii] > 0) & (dmat[ii] <= radius))[0]
        pick = within_r
        if pick.size < min_k:
            order = np.argsort(dmat[ii])[1:]
            pick = order[:max(k_fallback, min_k)]
        if pick.size > max_k:
            pick = pick[np.argsort(dmat[ii][pick])[:max_k]]

        neigh_idx = idxs[pick]
        if len(neigh_idx) < min_k:
            continue
        X = coords[neigh_idx]
        y = z[neigh_idx]
        a, b, c0 = fit_plane(X, y)
        x0, y0 = coords[gi]
        zhat0 = a * x0 + b * y0 + c0
        resid0 = z[gi] - zhat0
        res_neigh = y - (a * X[:, 0] + b * X[:, 1] + c0)

        med = float(np.median(res_neigh))
        mad = float(np.median(np.abs(res_neigh - med)))
        q1, q3 = np.percentile(res_neigh, [25, 75])
        iqr = float(q3 - q1)

        flags = []
        if mad > 0 and abs(resid0 - med) > mad_mult * mad:
            flags.append(">|MAD|")
        if iqr > 0 and (resid0 < q1 - iqr_mult * iqr or resid0 > q3 + iqr_mult * iqr):
            flags.append("Outside IQR")

        if flags:
            out.append({
                "Point ID": df.iloc[gi]["Point ID"],
                "Elevation": float(df.iloc[gi]["Elevation_num"]),
                "PredictedZ": float(zhat0),
                "Residual": float(resid0),
                "Neighbors": int(len(neigh_idx)),
                "Reason": "; ".join(flags),
            })

    return pd.DataFrame(out) if out else pd.DataFrame({"Info": ["No local elevation outliers found"]})

def _load_features_sheet(features_path: Path) -> pd.DataFrame:
    """Return the survey-features sheet with flexible sheet matching.

    Some user workbooks ship with the "Survey" sheet renamed (for example,
    "survey" or "Sheet1").  The previous implementation hard-failed in those
    scenarios.  Here we first attempt the canonical sheet name and then look for
    a case-insensitive match.  As a final fallback we use the first sheet in the
    workbook when it is unambiguous (only one sheet is present).
    """

    try:
        return pd.read_excel(features_path, sheet_name="Survey")
    except ValueError:
        xl = pd.ExcelFile(features_path)
        lower_map = {name.lower(): name for name in xl.sheet_names}
        if "survey" in lower_map:
            return xl.parse(lower_map["survey"])
        if len(xl.sheet_names) == 1:
            return xl.parse(xl.sheet_names[0])
        sheets = ", ".join(xl.sheet_names)
        raise ValueError(
            "Worksheet named 'Survey' not found. Available sheets: " + sheets
        )


def qc_pipeline(csv_path: Path, features_path: Path, output_dir: Path,
                radius=100.0, min_k=8, k_fallback=15, max_k=30, iqr_mult=3.0, mad_mult=6.0):
    """Runs the QC and returns all result DataFrames + metadata for reporting."""
    # Load features definitions
    feat = _load_features_sheet(features_path)
    feat = normalize_cols(feat)
    alpha_col = best_col(feat, "Alpha Code")
    atype_col = best_col(feat, "Attribute Type")
    feat["Alpha_base"] = feat[alpha_col].astype(str).map(strip_num_suffix)
    eligible_bases = set(feat.loc[feat[atype_col].astype(str).str.lower() != "do not include", "Alpha_base"])
    all_bases = set(feat["Alpha_base"])

    # Load survey CSV
    df = read_csv_any(csv_path)
    df = normalize_cols(df)
    # Canonical columns
    for want in ["Point ID", "Northing", "Easting", "Elevation", "Feature Code", "Attribute"]:
        actual = best_col(df, want)
        if actual != want:
            df.rename(columns={actual: want}, inplace=True)

    # Numeric fields
    df["Northing_num"] = pd.to_numeric(df["Northing"], errors="coerce")
    df["Easting_num"] = pd.to_numeric(df["Easting"], errors="coerce")
    df["Elevation_num"] = pd.to_numeric(df["Elevation"], errors="coerce")
    df["BaseCode"] = df["Feature Code"].astype(str).map(strip_num_suffix)

    # Non-numeric
    nn = []
    for c in ["Northing", "Easting", "Elevation"]:
        mask = df[c].notna() & df[c].astype(str).str.strip().ne("") & df[c + "_num"].isna()
        if mask.any():
            tmp = df.loc[mask, ["Point ID", c]].copy()
            tmp["Issue"] = f"Non-numeric {c}"
            nn.append(tmp)
    non_numeric_df = pd.concat(nn, ignore_index=True) if nn else pd.DataFrame({"Info": ["No issues found"]})

    # Missing
    miss = []
    for c in ["Point ID", "Northing_num", "Easting_num", "Elevation_num", "Feature Code"]:
        mask = df[c].isna() | (df[c].astype(str).str.strip() == "")
        if mask.any():
            tmp = df.loc[mask, ["Point ID", "Northing", "Easting", "Elevation", "Feature Code", "Attribute"]].copy()
            tmp["MissingField"] = c
            miss.append(tmp)
    missing_df = pd.concat(miss, ignore_index=True) if miss else pd.DataFrame({"Info": ["No issues found"]})

    # Duplicates
    dups = df[df["Point ID"].astype(str).duplicated(keep=False)].sort_values("Point ID")
    dups = dups[["Point ID", "Northing", "Easting", "Elevation", "Feature Code", "Attribute"]]
    if dups.empty:
        dups = pd.DataFrame({"Info": ["No duplicates"]})

    # Attribute format sanity (odd number of pipes)
    attr_issues = []
    for idx, raw in df["Attribute"].fillna("").astype(str).items():
        s = raw.strip()
        if s and s.count("|") % 2 == 1:
            attr_issues.append({
                "RowIndex": idx,
                "Point ID": df.at[idx, "Point ID"],
                "Attribute": raw,
                "Issue": "Odd number of '|' separators"
            })
    attr_issues_df = pd.DataFrame(attr_issues) if attr_issues else pd.DataFrame({"Info": ["No issues found"]})

    # Geometry warnings
    geom = []
    zero_mask = (df["Northing_num"] == 0) | (df["Easting_num"] == 0)
    big_mask = (df["Northing_num"].abs() > 1e7) | (df["Easting_num"].abs() > 1e7)
    if zero_mask.any():
        geom.append(df.loc[zero_mask, ["Point ID", "Northing", "Easting"]].assign(Issue="Zero coordinate"))
    if big_mask.any():
        geom.append(df.loc[big_mask, ["Point ID", "Northing", "Easting"]].assign(Issue="Coordinate magnitude unusually large"))
    geom_df = pd.concat(geom, ignore_index=True) if geom else pd.DataFrame({"Info": ["No issues found"]})

    # Unknown codes vs features
    unknown_mask = df["BaseCode"].ne("") & ~df["BaseCode"].isin(all_bases)
    unknown_codes = df.loc[unknown_mask, ["Point ID", "Feature Code"]].copy()
    if unknown_codes.empty:
        unknown_codes = pd.DataFrame({"Info": ["All feature bases recognized (Survey Features Excel)"]})

    # Linework events
    events = []
    for row_order, (i, row) in enumerate(df.iterrows()):
        fc_events = extract_line_events_field(
            row["Feature Code"] if pd.notna(row["Feature Code"]) else ""
        )
        at_events = extract_line_events_field(
            row["Attribute"] if pd.notna(row["Attribute"]) else ""
        )
        for base_val, line_id_val, ctrl_val in fc_events + at_events:
            events.append(
                {
                    "Row": i,
                    "RowOrder": row_order,
                    "Point ID": row["Point ID"],
                    "Base": base_val or "",
                    "LineID": "" if line_id_val is None else str(line_id_val or ""),
                    "Ctrl": ctrl_val,
                }
            )

    if events:
        events_df = pd.DataFrame(events)
        row_base_series = pd.Series(
            df.loc[events_df["Row"], "BaseCode"].values, index=events_df.index
        )
        events_df["Base"] = events_df["Base"].where(
            events_df["Base"].notna() & events_df["Base"].astype(str).str.strip().ne(""),
            row_base_series.fillna(""),
        )
        events_df["Base"] = events_df["Base"].fillna("")
        df["_PID_num"] = pd.to_numeric(df["Point ID"], errors="coerce")
        events_df = events_df.merge(
            df[["Point ID", "_PID_num"]]
            .reset_index()
            .rename(columns={"index": "Row"}),
            on=["Row", "Point ID"],
            how="left",
        )
        events_df["Ctrl"] = events_df["Ctrl"].str.upper()
        events_df = (
            events_df.sort_values(by=["_PID_num", "Point ID", "Row"])
            .reset_index(drop=True)
        )
        events_df.drop(columns=["_PID_num"], inplace=True)
        df.drop(columns=["_PID_num"], inplace=True)
    else:
        events_df = pd.DataFrame({"Info": ["No line events detected"]})

    if "Ctrl" in events_df.columns and not events_df.empty:
        state = {}
        for idx, ev in events_df.iterrows():
            base_val = (ev.get("Base") or "").strip()
            line_id_val = str(ev.get("LineID") or "")
            ctrl_val = str(ev.get("Ctrl") or "").upper()
            key = (base_val, line_id_val)
            if key not in state:
                state[key] = {
                    "open": False,
                    "pc_open": False,
                    "starts": 0,
                    "ends": 0,
                    "closes": 0,
                    "pcs": 0,
                    "pts": 0,
                    "issues": [],
                    "first_event_idx": idx,
                }
            s = state[key]

            if ctrl_val in CONTROL_START:
                if not s["open"]:
                    s["open"] = True
                else:
                    s["issues"].append("Start while already open")
                s["starts"] += 1
            elif ctrl_val in CONTROL_END:
                if s["open"]:
                    s["open"] = False
                    if s["pc_open"]:
                        s["issues"].append("Line closed with PC still open")
                        s["pc_open"] = False
                else:
                    s["issues"].append("Ended/closed with no prior start")
                s["ends"] += 1
            elif ctrl_val in CONTROL_CLOSE:
                if s["open"]:
                    s["open"] = False
                    if s["pc_open"]:
                        s["issues"].append("Line closed with PC still open")
                        s["pc_open"] = False
                else:
                    s["issues"].append("Ended/closed with no prior start")
                s["closes"] += 1
            elif ctrl_val in CONTROL_PC:
                if not s["open"]:
                    s["issues"].append("PC before start")
                elif s["pc_open"]:
                    s["issues"].append("PC while already in PC segment")
                else:
                    s["pc_open"] = True
                s["pcs"] += 1
            elif ctrl_val in CONTROL_PT:
                if s["pc_open"]:
                    s["pc_open"] = False
                else:
                    s["issues"].append("PT without prior PC")
                s["pts"] += 1

        summary_rows = []
        for key, s in sorted(state.items(), key=lambda kv: kv[1]["first_event_idx"]):
            base_key, line_id_key = key
            if s["open"]:
                s["issues"].append("Started but never ended/closed")
            if s["pc_open"]:
                s["issues"].append("PC without PT by file end")
            summary_rows.append(
                {
                    "Base": base_key,
                    "LineID": line_id_key,
                    "Starts": s["starts"],
                    "Ends": s["ends"],
                    "Closes": s["closes"],
                    "PCs": s["pcs"],
                    "PTs": s["pts"],
                    "Issues": "; ".join(s["issues"]) if s["issues"] else "OK",
                }
            )

        linework_df = (
            pd.DataFrame(summary_rows)
            if summary_rows
            else pd.DataFrame({"Info": ["No line events detected"]})
        )
    else:
        linework_df = pd.DataFrame({"Info": ["No line events detected"]})

    # Elevation outliers (terrain-eligible only)
    terrain_df = df[df["BaseCode"].isin(eligible_bases)].reset_index(drop=True)
    local_outliers = local_plane_outliers(
        terrain_df, radius=radius, min_k=min_k, k_fallback=k_fallback,
        max_k=max_k, iqr_mult=iqr_mult, mad_mult=mad_mult
    )

    # Build Summary counts
    def count_or_zero(d: pd.DataFrame) -> int:
        return 0 if "Info" in d.columns else len(d)

    summary = pd.DataFrame({
        "Category": [
            "Unknown Feature Codes", "Linework Issues", "Linework Events",
            "Non-numeric", "Missing Required", "Duplicate Point IDs",
            "Geometry Warnings", "Attribute Format Issues", "Local Elevation Outliers"
        ],
        "Count": [
            count_or_zero(unknown_codes),
            0 if "Info" in linework_df.columns else (linework_df["Issues"] != "OK").sum(),
            0 if "Info" in events_df.columns else len(events_df),
            count_or_zero(non_numeric_df),
            count_or_zero(missing_df),
            0 if "Info" in dups.columns else len(dups),
            count_or_zero(geom_df),
            count_or_zero(attr_issues_df),
            count_or_zero(local_outliers)
        ]
    })

    # Return everything (report function will decide how to write)
    return {
        "summary": summary,
        "unknown_codes": unknown_codes,
        "linework_issues": linework_df,
        "linework_events": events_df,
        "non_numeric": non_numeric_df,
        "missing": missing_df,
        "duplicates": dups,
        "geometry": geom_df,
        "attr_issues": attr_issues_df,
        "elev_outliers": local_outliers
    }

# ============ Reporting (Excel or HTML) ============
def write_excel_report(dfs: dict, output_path: Path):
    with pd.ExcelWriter(output_path, engine=XLSX_ENGINE) as w:
        # Summary
        dfs["summary"].to_excel(w, "Summary", index=False)

        def write_tab(name, key):
            df_ = dfs[key]
            if df_ is None or df_.empty:
                df_ = pd.DataFrame({"Info": ["No issues found"]})
            df_.to_excel(w, name, index=False)

        write_tab("Unknown Codes", "unknown_codes")
        write_tab("Line Continuity", "linework_issues")
        write_tab("Linework Events", "linework_events")
        write_tab("Non-numeric", "non_numeric")
        write_tab("Missing Required", "missing")
        write_tab("Duplicate Point IDs", "duplicates")
        write_tab("Geometry Warnings", "geometry")
        write_tab("Attribute Format Issues", "attr_issues")
        write_tab("Local Elevation Outliers", "elev_outliers")

    # Optional header styling
    if load_workbook and PatternFill and Font:
        try:
            wb = load_workbook(output_path)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for cell in ws[1]:
                    cell.fill = PatternFill(start_color='0F3320', end_color='0F3320', fill_type='solid')
                    cell.font = Font(color='FFFFFF', bold=True)
            wb.save(output_path)
        except Exception:
            pass

def _df_to_html(df: pd.DataFrame) -> str:
    if df is None or df.empty:
        df = pd.DataFrame({"Info": ["No issues found"]})
    # Add basic classes for CSS
    return df.to_html(index=False, classes="table table-zebra", border=0, escape=False)

def write_html_report(dfs: dict, output_path: Path, title="GPI Survey QC Report"):
    logo_data_uri = ""
    logo_path = Path(__file__).with_name("GPI-768x768.jpg")
    if logo_path.exists():
        try:
            encoded = base64.b64encode(logo_path.read_bytes()).decode("ascii")
            logo_data_uri = f"data:image/jpeg;base64,{encoded}"
        except Exception:
            logo_data_uri = ""

    css = f"""
    <style>
      :root {{
        --gpi-green: {GPI_GREEN};
        --gpi-med: {GPI_MED};
        --gpi-grey: {GPI_GREY};
        --gpi-hl: {GPI_HL};
        --bg: {BG};
      }}
      * {{ box-sizing: border-box; }}
      body {{
        margin: 0; font-family: Segoe UI, Roboto, Arial, sans-serif; background: var(--bg); color: #1a1a1a;
      }}
      header {{
        position: sticky; top: 0; z-index: 10; background: var(--gpi-green);
        color: #fff; padding: 16px 24px; box-shadow: 0 2px 8px rgba(0,0,0,.15);
        display: flex; align-items: center; gap: 16px;
      }}
      header img.logo {{ height: 48px; width: auto; border-radius: 8px; box-shadow: 0 2px 6px rgba(0,0,0,.25); }}
      h1 {{ margin: 0; font-size: 22px; }}
      .container {{ display: grid; grid-template-columns: 280px 1fr; gap: 20px; padding: 20px; }}
      nav {{
        background: #fff; border-radius: 12px; padding: 16px; box-shadow: 0 2px 12px rgba(0,0,0,.08);
        position: sticky; top: 76px; height: calc(100vh - 96px); overflow: auto;
      }}
      nav h2 {{ font-size: 14px; margin-top: 0; color: var(--gpi-med); letter-spacing: .3px; }}
      nav a {{
        display: flex; justify-content: space-between; align-items: center;
        padding: 10px 12px; margin: 6px 0; border-radius: 8px; text-decoration: none; color: #222;
      }}
      nav a:hover {{ background: #f1f5f2; }}
      .badge {{
        background: #eef7ea; color: #0d4b1f; padding: 2px 8px; border-radius: 10px; font-size: 12px;
      }}
      .ok {{ background: #eaf7ea; color: #0d4b1f; }}
      .warn {{ background: #fdecec; color: #7d0b0b; }}
      section {{
        background: #fff; border-radius: 12px; padding: 16px 16px 6px;
        box-shadow: 0 2px 12px rgba(0,0,0,.08); margin-bottom: 20px;
      }}
      section h3 {{ margin: 0 0 10px 0; color: var(--gpi-green); }}
      .table {{
        width: 100%; border-collapse: collapse; font-size: 13.5px; margin-bottom: 14px;
      }}
      .table thead th {{
        position: sticky; top: 0; background: var(--gpi-green); color: #fff; text-align: left; padding: 10px;
      }}
      .table td {{ padding: 8px 10px; border-bottom: 1px solid #eee; vertical-align: top; }}
      .table-zebra tbody tr:nth-child(odd) {{ background: #fafcfa; }}
      .muted {{ color: #666; font-size: 13px; }}
    </style>
    """

    # Build summary links
    summary_df = dfs["summary"].copy()
    sections = [
        ("Unknown Codes","unknown_codes"),
        ("Linework Issues","linework_issues"),
        ("Linework Events","linework_events"),
        ("Non-numeric","non_numeric"),
        ("Missing Required","missing"),
        ("Duplicate Point IDs","duplicates"),
        ("Geometry Warnings","geometry"),
        ("Attribute Format Issues","attr_issues"),
        ("Local Elevation Outliers","elev_outliers"),
    ]
    link_items = []
    for label, key in sections:
        row = summary_df[summary_df["Category"].str.contains(label.split()[0], case=False, regex=False)]
        count = (int(row["Count"].iloc[0]) if not row.empty else 0)
        badge_class = "ok" if count == 0 else "warn"
        link_items.append(f'<a href="#{key}"><span>{label}</span><span class="badge {badge_class}">{count}</span></a>')

    # Sections HTML
    def section_block(label, key, subtitle=""):
        df = dfs[key]
        table_html = _df_to_html(df)
        sub = f'<div class="muted">{subtitle}</div>' if subtitle else ""
        return f'''
          <section id="{key}">
            <h3>{label}</h3>
            {sub}
            {table_html}
          </section>
        '''

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8" />
<title>{title}</title>
{css}
</head>
<body>
<header>
  {f'<img src="{logo_data_uri}" alt="GPI logo" class="logo" />' if logo_data_uri else ''}
  <h1>{title}</h1>
</header>
<div class="container">
  <nav>
    <h2>Summary</h2>
    {''.join(link_items)}
  </nav>
  <main>
    {section_block("Unknown Codes", "unknown_codes")}
    {section_block("Linework Issues", "linework_issues",
      "Flags: 'Ended/closed with no prior start' or 'Started but never ended/closed'.")}
    {section_block("Linework Events", "linework_events",
      "START/END/CLOSE events tied to the flagged linework above.")}
    {section_block("Non-numeric", "non_numeric")}
    {section_block("Missing Required", "missing")}
    {section_block("Duplicate Point IDs", "duplicates")}
    {section_block("Geometry Warnings", "geometry")}
    {section_block("Attribute Format Issues", "attr_issues",
      "Checks for odd number of '|' separators in Attribute column.")}
    {section_block("Local Elevation Outliers", "elev_outliers",
      "Terrain-eligible codes only (Attribute Type != 'Do Not Include').")}
  </main>
</div>
</body>
</html>"""

    output_path.write_text(html, encoding="utf-8")

# ============ GUI ============
class SettingsDialog(tk.Toplevel):
    def __init__(self, master, current_format: str):
        super().__init__(master)
        self.title("Settings")
        self.configure(bg=BG)
        self.resizable(False, False)
        self.result = current_format  # 'excel' or 'html'

        tk.Label(self, text="Output format", bg=BG, fg=GPI_GREEN,
                 font=("Segoe UI", 11, "bold")).pack(anchor="w", padx=14, pady=(12, 4))

        self.var = tk.StringVar(value=current_format)
        frm = tk.Frame(self, bg=BG)
        frm.pack(anchor="w", padx=14, pady=(0, 10))
        tk.Radiobutton(frm, text="Excel (.xlsx)", variable=self.var, value="excel", bg=BG).pack(anchor="w")
        tk.Radiobutton(frm, text="HTML (.html)", variable=self.var, value="html", bg=BG).pack(anchor="w")

        btns = tk.Frame(self, bg=BG)
        btns.pack(fill="x", padx=14, pady=(6, 12))
        tk.Button(btns, text="Cancel", command=self._cancel).pack(side="right", padx=(6,0))
        tk.Button(btns, text="Save", bg=GPI_HL, command=self._save).pack(side="right")

        self.grab_set()
        self.protocol("WM_DELETE_WINDOW", self._cancel)

    def _save(self):
        self.result = self.var.get()
        self.destroy()

    def _cancel(self):
        self.result = None
        self.destroy()

BaseTk = TkinterDnD.Tk if TkinterDnD else tk.Tk


class App(BaseTk):
    def __init__(self):
        super().__init__()
        self.title("GPI Survey QC Tool")
        self.geometry("900x700")
        self.configure(bg=BG)

        self.log_q = queue.Queue()
        self.output_format = "excel"  # or "html"
        self.logo_photo = None

        if Image and ImageTk:
            logo_path = Path(__file__).with_name("GPI-768x768.jpg")
            if logo_path.exists():
                try:
                    img = Image.open(logo_path)
                    img.thumbnail((48, 48), getattr(Image, "LANCZOS", Image.BICUBIC))
                    self.logo_photo = ImageTk.PhotoImage(img)
                except Exception:
                    self.logo_photo = None

        # Header with gear on right
        header = tk.Frame(self, bg=GPI_GREEN)
        header.pack(fill=tk.X)
        if self.logo_photo is not None:
            tk.Label(header, image=self.logo_photo, bg=GPI_GREEN).pack(side="left", padx=(16, 8), pady=6)
        tk.Label(header, text="GPI Survey QC Tool", bg=GPI_GREEN, fg="white",
                 font=("Segoe UI", 16, "bold"), pady=10).pack(side="left", padx=(0, 16))
        tk.Button(header, text="⚙", bg=GPI_GREEN, fg="white", bd=0, font=("Segoe UI", 14, "bold"),
                  activebackground=GPI_HL, activeforeground="black",
                  command=self._open_settings).pack(side="right", padx=12, pady=6)

        # Paths
        top = tk.Frame(self, bg=BG)
        top.pack(fill=tk.X, padx=12, pady=10)

        tk.Label(top, text="Survey CSV:", bg=BG, fg=GPI_GREEN).grid(row=0, column=0, sticky="w")
        self.csv_entry = tk.Entry(top, width=64)
        self.csv_entry.grid(row=0, column=1, padx=6)
        self._register_file_drop(self.csv_entry)
        tk.Button(top, text="Browse", bg=GPI_HL, command=self._pick_csv).grid(row=0, column=2)

        tk.Label(top, text="Survey Features Excel (Survey sheet):", bg=BG, fg=GPI_GREEN).grid(row=1, column=0, sticky="w")
        self.feat_entry = tk.Entry(top, width=64)
        self.feat_entry.grid(row=1, column=1, padx=6)
        self._register_file_drop(self.feat_entry)
        tk.Button(top, text="Browse", bg=GPI_HL, command=self._pick_feat).grid(row=1, column=2)

        tk.Label(top, text="Output Folder:", bg=BG, fg=GPI_GREEN).grid(row=2, column=0, sticky="w")
        self.out_entry = tk.Entry(top, width=64)
        self.out_entry.grid(row=2, column=1, padx=6)
        self._register_file_drop(self.out_entry)
        tk.Button(top, text="Browse", bg=GPI_HL, command=self._pick_out).grid(row=2, column=2)

        # Params
        params = tk.Frame(self, bg=BG)
        params.pack(fill=tk.X, padx=12, pady=(0, 10))
        tk.Label(params, text="Local elevation parameters", bg=BG, fg="#333",
                 font=("Segoe UI", 10, "italic")).grid(row=0, column=0, columnspan=6, sticky="w", pady=(0, 6))

        self.radius = tk.DoubleVar(value=100.0)
        self.min_k = tk.IntVar(value=8)
        self.k_fb = tk.IntVar(value=15)
        self.max_k = tk.IntVar(value=30)
        self.iqr_mult = tk.DoubleVar(value=3.0)
        self.mad_mult = tk.DoubleVar(value=6.0)

        def add_param(r, c, label, var):
            tk.Label(params, text=label, bg=BG, fg=GPI_GREEN).grid(row=r, column=c, sticky="e", padx=4)
            tk.Entry(params, textvariable=var, width=8).grid(row=r, column=c+1, sticky="w")

        add_param(1, 0, "Radius", self.radius)
        add_param(1, 2, "min_k", self.min_k)
        add_param(1, 4, "k_fallback", self.k_fb)
        add_param(2, 0, "max_k", self.max_k)
        add_param(2, 2, "IQR mult", self.iqr_mult)
        add_param(2, 4, "MAD mult", self.mad_mult)

        # >>> Added explanatory help text (no logic changes) <<<
        help_text = (
            "Elevation QC parameters (recommendations):\n"
            "• Radius — neighborhood search distance in your horizontal units. "
            "Typical 50–150 (default 100). Larger = smoother, fewer flags; smaller = more sensitive.\n"
            "• min_k — minimum neighbors to fit the plane. Typical 6–12 (default 8).\n"
            "• k_fallback — neighbors used if not enough points found within Radius. Typical 12–20 (default 15).\n"
            "• max_k — hard cap on neighbors for the fit. Typical 20–50 (default 30).\n"
            "• IQR mult — robust spread threshold (Q1–Q3). Typical 2.5–3.5 (default 3.0).\n"
            "• MAD mult — robust outlier threshold. Typical 5–8 (default 6.0).\n"
            "Tip: If you see too many outliers, increase Radius or thresholds; if too few, decrease them."
        )
        tk.Label(
            params, text=help_text, justify="left", anchor="w",
            bg=BG, fg="#4a4a4a", wraplength=760
        ).grid(row=3, column=0, columnspan=6, sticky="w", pady=(8, 0))

        # Run button
        self.run_btn = tk.Button(self, text="Run QC", bg=GPI_MED, fg="white",
                                 font=("Segoe UI", 12, "bold"), command=self._start_thread)
        self.run_btn.pack(pady=10)

        # Log
        self.log = scrolledtext.ScrolledText(self, height=14, bg="white", fg="black")
        self.log.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)

        # Status bar
        self.status = tk.StringVar(value=f"Ready — Output: {self.output_format.upper()}")
        tk.Label(self, textvariable=self.status, bg=GPI_GREEN, fg="white", anchor="w").pack(fill=tk.X)

        self.after(150, self._pump)

    def _register_file_drop(self, entry):
        if not (TkinterDnD and DND_FILES):
            return
        try:
            entry.drop_target_register(DND_FILES)
            entry.dnd_bind("<<Drop>>", lambda e, ent=entry: self._on_drop_files(e, ent))
        except Exception:
            # If TkDnD is not fully available on the platform, continue without drag-and-drop.
            pass

    def _on_drop_files(self, event, entry):
        data = getattr(event, "data", "")
        if not data:
            return
        try:
            paths = entry.tk.splitlist(data)
        except Exception:
            paths = [data]
        if not paths:
            return
        first = str(paths[0]).strip()
        if first.startswith("{") and first.endswith("}"):
            first = first[1:-1]
        entry.delete(0, "end")
        entry.insert(0, first)
        entry.focus_set()

    # Settings (gear)
    def _open_settings(self):
        dlg = SettingsDialog(self, self.output_format)
        self.wait_window(dlg)
        if dlg.result in ("excel", "html"):
            self.output_format = dlg.result
            self.status.set(f"Ready — Output: {self.output_format.upper()}")

    # File pickers
    def _pick_csv(self):
        p = filedialog.askopenfilename(title="Select Survey CSV", filetypes=[("CSV files", "*.csv"), ("All files", "*.*")])
        if p:
            self.csv_entry.delete(0, "end"); self.csv_entry.insert(0, p)

    def _pick_feat(self):
        p = filedialog.askopenfilename(title="Select Survey Features Excel", filetypes=[("Excel files", "*.xlsx;*.xls")])
        if p:
            self.feat_entry.delete(0, "end"); self.feat_entry.insert(0, p)

    def _pick_out(self):
        p = filedialog.askdirectory(title="Select Output Folder")
        if p:
            self.out_entry.delete(0, "end"); self.out_entry.insert(0, p)

    # Threading
    def _start_thread(self):
        if not self.csv_entry.get() or not self.feat_entry.get() or not self.out_entry.get():
            messagebox.showerror("Missing info", "Please select a Survey CSV, Survey Features Excel, and an output folder.")
            return
        self.run_btn.config(state="disabled")
        self.status.set(f"Running… (Output: {self.output_format.upper()})")
        t = threading.Thread(target=self._run_qc, daemon=True)
        t.start()

    def _run_qc(self):
        csv = feat = outdir = None
        try:
            csv = Path(self.csv_entry.get())
            feat = Path(self.feat_entry.get())
            outdir = Path(self.out_entry.get())
            self.log_q.put(f"Survey: {csv}")
            self.log_q.put(f"Features: {feat}")
            self.log_q.put(f"Output folder: {outdir}")
            # Run QC
            results = qc_pipeline(
                csv, feat, outdir,
                radius=float(self.radius.get()),
                min_k=int(self.min_k.get()),
                k_fallback=int(self.k_fb.get()),
                max_k=int(self.max_k.get()),
                iqr_mult=float(self.iqr_mult.get()),
                mad_mult=float(self.mad_mult.get())
            )
            # Write selected format
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            if self.output_format == "excel":
                out = outdir / f"qc_report_{ts}.xlsx"
                write_excel_report(results, out)
            else:
                out = outdir / f"qc_report_{ts}.html"
                write_html_report(results, out, title="GPI Survey QC Report")
            self.log_q.put(f"Report generated: {out}")
            self.status.set("Done")
            messagebox.showinfo("QC Complete", f"Report saved to:\n{out}")
        except Exception as e:
            error_code = _generate_error_code()
            context = {
                "Survey CSV": str(csv) if csv else "",
                "Survey Features": str(feat) if feat else "",
                "Output folder": str(outdir) if outdir else "",
            }
            log_path = _write_crash_log(
                error_code,
                e,
                context=context,
                preferred_dir=outdir if outdir else None,
            )
            self.status.set("Error")
            self.log_q.put(f"ERROR [{error_code}]: {e}")
            if log_path:
                self.log_q.put(f"Crash log saved to: {log_path}")
            else:
                self.log_q.put("Unable to save crash log.")

            user_message = [
                "An unexpected error occurred while running QC.",
                f"Error code: {error_code}",
            ]
            if log_path:
                user_message.append(f"Details were saved to:\n{log_path}")
            else:
                user_message.append("The crash log could not be written. See console for details.")
            user_message.append(f"Original error:\n{e}")
            messagebox.showerror("QC Failed", "\n\n".join(user_message))
        finally:
            self.run_btn.config(state="normal")

    # Logger pump
    def _pump(self):
        try:
            while not self.log_q.empty():
                msg = self.log_q.get_nowait()
                self.log.insert("end", msg + "\n")
                self.log.see("end")
        except queue.Empty:
            pass
        self.after(150, self._pump)

def main():
    App().mainloop()


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:  # pragma: no cover - fatal fallback
        error_code = _generate_error_code()
        log_path = _write_crash_log(error_code, exc)
        err_lines = [
            "A fatal error occurred while starting the GPI Survey QC Tool.",
            f"Error code: {error_code}",
        ]
        if log_path:
            err_lines.append(f"Crash log saved to: {log_path}")
        else:
            err_lines.append("Additionally, the crash log could not be written.")
        err_lines.append("See console for more details.")
        sys.stderr.write("\n".join(err_lines) + "\n")
        traceback.print_exception(type(exc), exc, exc.__traceback__)
